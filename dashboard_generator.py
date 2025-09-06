"""
Dashboard generator for YMCA volunteer data
Creates visualizations and exports for PowerBI/Excel
"""

import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import json
from datetime import datetime
from typing import Dict, List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class DashboardGenerator:
    """Generate dashboards and visualizations for volunteer data"""
    
    def __init__(self):
        """Initialize dashboard generator"""
        self.colors = {
            'primary': '#003f7f',  # YMCA Blue
            'secondary': '#ed1c24',  # YMCA Red
            'accent': '#00a859',  # Green
            'warning': '#ffc627',  # Yellow
            'info': '#0095da'  # Light Blue
        }
    
    def create_hours_by_branch_chart(self, branch_stats: Dict) -> go.Figure:
        """
        Create bar chart showing volunteer hours by branch
        
        Args:
            branch_stats: Dictionary with branch statistics
            
        Returns:
            Plotly figure
        """
        hours_data = branch_stats.get('hours_by_branch', {})
        
        if not hours_data:
            return go.Figure().add_annotation(
                text="No data available",
                showarrow=False
            )
        
        branches = list(hours_data.keys())
        hours = list(hours_data.values())
        
        fig = go.Figure(data=[
            go.Bar(
                x=branches,
                y=hours,
                marker_color=self.colors['primary'],
                text=[f"{h:,.0f}" for h in hours],
                textposition='auto'
            )
        ])
        
        fig.update_layout(
            title="Volunteer Hours by Branch",
            xaxis_title="Branch",
            yaxis_title="Hours",
            showlegend=False,
            height=400
        )
        
        return fig
    
    def create_volunteers_by_branch_chart(self, branch_stats: Dict) -> go.Figure:
        """
        Create bar chart showing active volunteers by branch
        
        Args:
            branch_stats: Dictionary with branch statistics
            
        Returns:
            Plotly figure
        """
        volunteer_data = branch_stats.get('volunteers_by_branch', {})
        
        if not volunteer_data:
            return go.Figure().add_annotation(
                text="No data available",
                showarrow=False
            )
        
        branches = list(volunteer_data.keys())
        volunteers = list(volunteer_data.values())
        
        fig = go.Figure(data=[
            go.Bar(
                x=branches,
                y=volunteers,
                marker_color=self.colors['secondary'],
                text=volunteers,
                textposition='auto'
            )
        ])
        
        fig.update_layout(
            title="Active Volunteers by Branch",
            xaxis_title="Branch",
            yaxis_title="Number of Volunteers",
            showlegend=False,
            height=400
        )
        
        return fig
    
    def create_milestone_progress_chart(self, milestones: Dict) -> go.Figure:
        """
        Create chart showing milestone achievements
        
        Args:
            milestones: Dictionary with milestone data
            
        Returns:
            Plotly figure
        """
        achieved = milestones.get('achieved', [])
        
        if not achieved:
            return go.Figure().add_annotation(
                text="No milestone achievements yet",
                showarrow=False
            )
        
        # Count achievements by milestone level
        milestone_counts = {}
        for record in achieved:
            award = record['award_name']
            milestone_counts[award] = milestone_counts.get(award, 0) + 1
        
        # Order milestones
        ordered_milestones = [
            "First Impact",
            "Service Star",
            "Commitment Champion",
            "Passion In Action Award",
            "Guiding Light Award"
        ]
        
        awards = []
        counts = []
        for milestone in ordered_milestones:
            if milestone in milestone_counts:
                awards.append(milestone)
                counts.append(milestone_counts[milestone])
        
        fig = go.Figure(data=[
            go.Bar(
                x=awards,
                y=counts,
                marker_color=self.colors['accent'],
                text=counts,
                textposition='auto'
            )
        ])
        
        fig.update_layout(
            title="Volunteer Milestone Achievements",
            xaxis_title="Milestone Level",
            yaxis_title="Number of Volunteers",
            showlegend=False,
            height=400
        )
        
        return fig
    
    def create_project_category_pie(self, project_stats: Dict) -> go.Figure:
        """
        Create pie chart showing hours distribution by project category
        
        Args:
            project_stats: Dictionary with project statistics
            
        Returns:
            Plotly figure
        """
        hours_data = project_stats.get('hours_by_category', {})
        
        if not hours_data:
            return go.Figure().add_annotation(
                text="No data available",
                showarrow=False
            )
        
        categories = list(hours_data.keys())
        hours = list(hours_data.values())
        
        fig = go.Figure(data=[
            go.Pie(
                labels=categories,
                values=hours,
                hole=0.3,
                marker=dict(
                    colors=px.colors.qualitative.Set3[:len(categories)]
                )
            )
        ])
        
        fig.update_layout(
            title="Volunteer Hours by Project Category",
            height=400
        )
        
        return fig
    
    def create_dashboard_html(self, report: Dict) -> str:
        """
        Create complete HTML dashboard
        
        Args:
            report: Complete report dictionary
            
        Returns:
            HTML string
        """
        # Create charts
        branch_hours_fig = self.create_hours_by_branch_chart(report['branch_stats'])
        branch_volunteers_fig = self.create_volunteers_by_branch_chart(report['branch_stats'])
        milestone_fig = self.create_milestone_progress_chart(report['milestones'])
        project_pie_fig = self.create_project_category_pie(report['project_stats'])
        
        # Convert to HTML
        branch_hours_html = branch_hours_fig.to_html(include_plotlyjs=False, div_id="branch_hours")
        branch_volunteers_html = branch_volunteers_fig.to_html(include_plotlyjs=False, div_id="branch_volunteers")
        milestone_html = milestone_fig.to_html(include_plotlyjs=False, div_id="milestones")
        project_pie_html = project_pie_fig.to_html(include_plotlyjs=False, div_id="project_pie")
        
        # Create summary cards HTML
        summary = report['summary']
        summary_html = f"""
        <div class="row">
            <div class="col-md-3">
                <div class="card text-center">
                    <div class="card-body">
                        <h5 class="card-title">Total Hours</h5>
                        <h2 class="text-primary">{summary['total_hours']:,.0f}</h2>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card text-center">
                    <div class="card-body">
                        <h5 class="card-title">Active Volunteers</h5>
                        <h2 class="text-success">{summary['total_volunteers']}</h2>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card text-center">
                    <div class="card-body">
                        <h5 class="card-title">Projects</h5>
                        <h2 class="text-info">{summary['unique_projects']}</h2>
                    </div>
                </div>
            </div>
            <div class="col-md-3">
                <div class="card text-center">
                    <div class="card-body">
                        <h5 class="card-title">Branches</h5>
                        <h2 class="text-warning">{summary['unique_branches']}</h2>
                    </div>
                </div>
            </div>
        </div>
        """
        
        # Complete HTML document
        html = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>YMCA Volunteer Dashboard - {report['report_month']}</title>
            <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
            <script src="https://cdn.plot.ly/plotly-latest.min.js"></script>
            <style>
                body {{
                    font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                    background-color: #f8f9fa;
                }}
                .navbar {{
                    background-color: #003f7f !important;
                }}
                .card {{
                    margin-bottom: 20px;
                    border: none;
                    box-shadow: 0 0 10px rgba(0,0,0,0.1);
                }}
                .card-body {{
                    padding: 20px;
                }}
                h1, h2 {{
                    color: #003f7f;
                }}
                .text-primary {{
                    color: #003f7f !important;
                }}
                .text-success {{
                    color: #00a859 !important;
                }}
                .text-info {{
                    color: #0095da !important;
                }}
                .text-warning {{
                    color: #ffc627 !important;
                }}
            </style>
        </head>
        <body>
            <nav class="navbar navbar-dark">
                <div class="container-fluid">
                    <span class="navbar-brand mb-0 h1">YMCA of Greater Cincinnati - Volunteer Dashboard</span>
                </div>
            </nav>
            
            <div class="container-fluid mt-4">
                <h1>{report['report_month']} Volunteer Report</h1>
                <p class="text-muted">Generated: {report['data_generated']}</p>
                
                {summary_html}
                
                <div class="row mt-4">
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-body">
                                {branch_hours_html}
                            </div>
                        </div>
                    </div>
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-body">
                                {branch_volunteers_html}
                            </div>
                        </div>
                    </div>
                </div>
                
                <div class="row">
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-body">
                                {project_pie_html}
                            </div>
                        </div>
                    </div>
                    <div class="col-md-6">
                        <div class="card">
                            <div class="card-body">
                                {milestone_html}
                            </div>
                        </div>
                    </div>
                </div>
            </div>
        </body>
        </html>
        """
        
        return html
    
    def export_to_excel(self, report: Dict, filename: str = None):
        """
        Export report data to Excel with formatting
        
        Args:
            report: Complete report dictionary
            filename: Output filename (optional)
        """
        if filename is None:
            filename = f"volunteer_report_{report['report_month'].replace(' ', '_')}.xlsx"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Add headers and data
        ws_summary['A1'] = f"YMCA Volunteer Report - {report['report_month']}"
        ws_summary['A1'].font = Font(size=16, bold=True)
        
        ws_summary['A3'] = "Summary Statistics"
        ws_summary['A3'].font = Font(size=14, bold=True)
        
        row = 5
        for key, value in report['summary'].items():
            ws_summary[f'A{row}'] = key.replace('_', ' ').title()
            ws_summary[f'B{row}'] = value
            row += 1
        
        # Branch statistics sheet
        ws_branch = wb.create_sheet("Branch Statistics")
        branch_data = []
        
        hours_by_branch = report['branch_stats'].get('hours_by_branch', {})
        volunteers_by_branch = report['branch_stats'].get('volunteers_by_branch', {})
        
        for branch in set(list(hours_by_branch.keys()) + list(volunteers_by_branch.keys())):
            branch_data.append({
                'Branch': branch,
                'Total Hours': hours_by_branch.get(branch, 0),
                'Active Volunteers': volunteers_by_branch.get(branch, 0)
            })
        
        if branch_data:
            df_branch = pd.DataFrame(branch_data)
            for r_idx, row in enumerate(dataframe_to_rows(df_branch, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_branch.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        ws_branch.cell(row=r_idx, column=c_idx).font = Font(bold=True)
        
        # Milestone achievements sheet
        ws_milestones = wb.create_sheet("Milestones")
        achieved = report['milestones'].get('achieved', [])
        
        if achieved:
            df_milestones = pd.DataFrame(achieved)
            for r_idx, row in enumerate(dataframe_to_rows(df_milestones, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws_milestones.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        ws_milestones.cell(row=r_idx, column=c_idx).font = Font(bold=True)
        
        # Save workbook
        wb.save(filename)
        print(f"Excel report saved to {filename}")
        
        return filename


def generate_dashboard(report_file: str = None, report_data: Dict = None):
    """
    Generate dashboard from report data
    
    Args:
        report_file: Path to JSON report file
        report_data: Report dictionary (if not loading from file)
    """
    generator = DashboardGenerator()
    
    # Load report data
    if report_data is None:
        with open(report_file, 'r') as f:
            report_data = json.load(f)
    
    # Generate HTML dashboard
    html = generator.create_dashboard_html(report_data)
    html_filename = f"dashboard_{report_data['report_month'].replace(' ', '_')}.html"
    
    with open(html_filename, 'w') as f:
        f.write(html)
    
    print(f"HTML dashboard saved to {html_filename}")
    
    # Generate Excel report
    excel_filename = generator.export_to_excel(report_data)
    
    return html_filename, excel_filename


if __name__ == "__main__":
    # Test with sample data
    print("Dashboard Generator initialized")